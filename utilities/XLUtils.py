import openpyxl


def getRowCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.max_row


def getColumnCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return (sheet.max_column)


def readData_multiple(file, sheetName, rownum, endRowNum, columnnum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    values = []
    for rownum in range(rownum, endRowNum + 1):
        values.append(sheet.cell(row=rownum, column=columnnum).value)
    return values


def readData(file, sheetName, rownum, columnnum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(row=rownum, column=columnnum).value


def writeData(file, sheetName, rownum, columnnum, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(row=rownum, column=columnnum).value = data
    workbook.save(file)
